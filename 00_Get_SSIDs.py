import requests
import json
import pandas as pd
import numpy as np
import os
import time
# import sys
import logging
import openpyxl


time_str = time.strftime('%m%d%Y_%H%M%S')
print(time_str)
logging.basicConfig(filename=f"LOG_file_{time_str}.log",
					format='%(asctime)s %(message)s',
					filemode='w')
logger = logging.getLogger()
# logger.setLevel(logging.INFO)
# Uncomment line below if need debugging in Log_file
logger.setLevel(logging.DEBUG)

# old_stdout = sys.stdout
# log_file = open("message.log","w")
# sys.stdout = log_file

df = pd.read_excel("Config_File_for_GET.xlsx", 'Step 0', skiprows=0)
df2 = df.replace(np.nan, '', regex=True)
data = []

for index, rows in df2.iterrows():
    data.append({rows[0]: rows[1]})

API_VALUE = data[0]['API Key']
ORG_ID = data[1]['Organization ID']

# NEW Helping def. Used for Step 1 and also Step 2. Get all Network with response in json
def getAllNetworks():
    url = f"https://api.meraki.com/api/v1/organizations/{ORG_ID}/networks"
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "X-Cisco-Meraki-API-Key": API_VALUE
    }
    payload = None
    response = requests.request('GET', url, headers=headers, data=payload)
    parsed = json.loads(response.text)
    apiPrintLogger = (json.dumps(parsed, indent=4, sort_keys=True))
    if response.status_code != 200:
        print(f'Error with status code {response.status_code}, text: {response.text}')
        logger.info(f'Error with status code {response.status_code}, text: {response.text}')
        logger.info("API RESPONSE\n")
        logger.info(apiPrintLogger)
        return []
    logger.info("API RESPONSE\n")
    logger.info(apiPrintLogger)
    # print(json.dumps(parsed, indent=4, sort_keys=True))
    return response.json()

# NEW Helping def. Used for Step 2. Get all SSIDs per network ID with response in json
def getSSIDs(net_id):
    url = f"https://api.meraki.com/api/v1/networks/{net_id}/wireless/ssids"
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "X-Cisco-Meraki-API-Key": API_VALUE
    }
    payload = None
    response = requests.request('GET', url, headers=headers, data=payload)
    parsed = json.loads(response.text)
    apiPrintLogger = (json.dumps(parsed, indent=4, sort_keys=True))
    if response.status_code != 200:
        print(f'Error with status code {response.status_code}, text: {response.text}')
        logger.info(f'Error with status code {response.status_code}, text: {response.text}')
        logger.info("API RESPONSE\n")
        logger.info(apiPrintLogger)
        return []
    logger.info("API RESPONSE\n")
    logger.info(apiPrintLogger)
    return response.json()

# NEW Step 1. See The List of Networks for ORG ID
def getNetworks():
    print(f"Network IDs for Organization {ORG_ID}: ")
    logger.info(f"Network IDs for Organization {ORG_ID}: ")
    networks = getAllNetworks()
    for net in networks:
        net_id = net['id']
        net_name = net['name']
        print(f'    Network ID: {net_id}, Name: {net_name}')
        logger.info(f'    Network ID: {net_id}, Name: {net_name}')

def getSSIDsTo_Excel():
    column_names = ['SSID_number', 'SSID_name', 'Enabled', 'Auth_Mode', 'Encrypt_Mode', 'wpa_Encrypt', 'Mode', 'Vlan_Tag', 'Min_Bitrate', 'Visible']
    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Remove the default sheet created
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create first sheet and name it with network name
    networks = getAllNetworks()
    for net in networks:
        net_id = net['id']
        net_name = net['name']
        print(f'    Creating Sheet for Network ID: {net_id}, Name: {net_name}')
        logger.info(f'    Creating Sheet for Network ID: {net_id}, Name: {net_name}')
        wb.create_sheet(net_name)

        # Add columns to the sheet based on column_names
        for col_idx, column_name in enumerate(column_names, start=1):
            sheet = wb[net_name]
            sheet.cell(row=1, column=col_idx, value=column_name)

        # Get SSIDs info
        SSIDs = getSSIDs(net_id)
        column_a_index = None
        for SSID in SSIDs:
            SSID_name = (SSID['name'])
            Enabled = SSID['enabled']
            SSID_number = str(SSID['number'])
            Auth_Mode = SSID['authMode']
            if 'encryptionMode' not in SSID:
                Encrypt_Mode = 'N/A'
            else:
                Encrypt_Mode = SSID['encryptionMode']
            if 'wpaEncryptionModee' not in SSID:
                wpa_Encrypt = 'N/A'
            else:
                wpa_Encrypt = SSID['wpaEncryptionMode']
            Mode = SSID['ipAssignmentMode']
            if 'defaultVlanId' not in SSID:
                Vlan_Tag = 'N/A'
            else:
                Vlan_Tag = SSID['defaultVlanId']
            Min_Bitrate = SSID['minBitrate']
            Visible = SSID['visible']

            print(f'SSID Info:\n    {SSID_number}\n    {SSID_name}\n    {Enabled}\n    {Auth_Mode}\n    {Encrypt_Mode}\n    {wpa_Encrypt}\n    {Mode}\n    {Vlan_Tag}\n    {Min_Bitrate}\n    {Visible}\n')
            for cell in sheet[1]:
                if cell.value == 'SSID_number':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                # Find the next available row in SSID_number column
                next_row = 2  # Start from the second row (assuming the first row is for headers)
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = SSID_number
                print(f"'{SSID_number}' added to column 'SSID_number' in row {next_row}")
                logger.info(f"'{SSID_number}' added to column 'SSID_number' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'SSID_name':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = SSID_name
                print(f"'{SSID_name}' added to column 'SSID_name' in row {next_row}")
                logger.info(f"'{SSID_name}' added to column 'SSID_name' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Enabled':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Enabled
                print(f"'{Enabled}' added to column 'Enabled' in row {next_row}")
                logger.info(f"'{Enabled}' added to column 'Enabled' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Auth_Mode':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Auth_Mode
                print(f"'{Auth_Mode}' added to column 'Auth_Mode' in row {next_row}")
                logger.info(f"'{Auth_Mode}' added to column 'Auth_Mode' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Encrypt_Mode':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Encrypt_Mode
                print(f"'{Encrypt_Mode}' added to column 'Encrypt_Mode' in row {next_row}")
                logger.info(f"'{Encrypt_Mode}' added to column 'Encrypt_Mode' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'wpa_Encrypt':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = wpa_Encrypt
                print(f"'{wpa_Encrypt}' added to column 'wpa_Encrypt' in row {next_row}")
                logger.info(f"'{wpa_Encrypt}' added to column 'wpa_Encrypt' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Mode':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Mode
                print(f"'{Mode}' added to column 'Mode' in row {next_row}")
                logger.info(f"'{Mode}' added to column 'Mode' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Vlan_Tag':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Vlan_Tag
                print(f"'{Vlan_Tag}' added to column 'Vlan_Tag' in row {next_row}")
                logger.info(f"'{Vlan_Tag}' added to column 'Vlan_Tag' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Min_Bitrate':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Min_Bitrate
                print(f"'{Min_Bitrate}' added to column 'Min_Bitrate' in row {next_row}")
                logger.info(f"'{Min_Bitrate}' added to column 'Min_Bitrate' in row {next_row}")
            for cell in sheet[1]:
                if cell.value == 'Visible':
                    column_a_index = cell.column_letter
                    break
            if column_a_index is not None:
                while sheet[f"{column_a_index}{next_row}"].value:
                    next_row += 1
                sheet[f"{column_a_index}{next_row}"] = Visible
                print(f"'{Visible}' added to column 'Visible' in row {next_row}")
                logger.info(f"'{Visible}' added to column 'Visible' in row {next_row}")

    # Save the workbook to a file
    ExcelFilename = f'Get_SSIDs_report_{time_str}.xlsx'
    wb.save(ExcelFilename)
    # Logger
    logger.info(f"Excel file created:\n    {ExcelFilename}\nwith following sheets:\n    {net_id}\nwith following "
                f"columns:\n    {column_names}")

if __name__ == "__main__":

    main_menu = True
    while main_menu:
        os.system('clear')
        print("=====================================================")
        print("=    Meraki Get Info Tool    =")
        print("=====================================================\n\n")
        print("[1] Get all Network ID and Names for ORG ID")
        print("[2] Get SSIDs. Get all SSIDs info for all Networks in Org ID")
        print("*[3] Not Ready yet")
        while True:
            step = input("Select a number [1-2]: ")
            logger.info(f'               SELECTED STEP: {step}')
            if step == "1":
                getNetworks()
                input("\nPress Enter to continue:")
                break
            if step == "2":
                getSSIDsTo_Excel()
                input("\nPress Enter to continue:")
                break
            if step == "3":
                print("This step is not ready yet")
                input("\nPress Enter to continue:")
                break

#sys.stdout = old_stdout
#log_file.close()